import pandas as pd
import openpyxl

def analyze_excel_structure(filename):
    """Analyze the structure of an existing Excel file"""
    print(f"\nAnalyzing {filename}:")
    print("=" * 50)
    
    try:
        # Load workbook to see sheet names
        wb = openpyxl.load_workbook(filename)
        print(f"Sheet names: {wb.sheetnames}")
        
        # Read each sheet
        for sheet_name in wb.sheetnames:
            print(f"\n--- Sheet: {sheet_name} ---")
            df = pd.read_excel(filename, sheet_name=sheet_name)
            print(f"Shape: {df.shape}")
            print(f"Columns: {list(df.columns)}")
            print("First few rows:")
            print(df.head())
            
    except Exception as e:
        print(f"Error reading {filename}: {e}")

# Analyze existing files
files_to_analyze = ['AAPL.xlsx', 'AAPL_stock_data.xlsx', 'MSFT_stock_data.xlsx', 'TSLA_stock_data.xlsx']

for file in files_to_analyze:
    try:
        analyze_excel_structure(file)
    except FileNotFoundError:
        print(f"File {file} not found")
    except Exception as e:
        print(f"Error with {file}: {e}")